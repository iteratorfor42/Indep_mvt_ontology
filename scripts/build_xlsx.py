"""
CSV(data/independence_ontology_factcheck.csv)를 읽어 서식이 적용된
data/independence_ontology_factcheck.xlsx 를 생성하는 스크립트.

사용법:
1. VSCode에서 data/independence_ontology_factcheck.csv 를 열어 내용 수정
2. 터미널에서 실행:  python scripts/build_xlsx.py
"""

import csv
import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# 현재 스크립트(scripts/build_xlsx.py) 폴더 기준으로 프로젝트 루트 및 data 폴더 경로 설정
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
PROJECT_ROOT = os.path.dirname(SCRIPT_DIR)
DATA_DIR = os.path.join(PROJECT_ROOT, "data")

CSV_PATH = os.path.join(DATA_DIR, "independence_ontology_factcheck.csv")
XLSX_PATH = os.path.join(DATA_DIR, "independence_ontology_factcheck.xlsx")

# ---- 1. CSV 읽기 ----
with open(CSV_PATH, encoding="utf-8-sig") as f:
    reader = csv.reader(f)
    all_rows = list(reader)

headers = all_rows[0]
rows = all_rows[1:]
last_row = len(rows) + 1  # 헤더 포함 마지막 데이터 행 번호

# ---- 2. 워크북 생성 ----
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "사실관계 목록"

ws.append(headers)
for r in rows:
    ws.append(r)

header_fill = PatternFill(start_color="2C3E66", end_color="2C3E66", fill_type="solid")
header_font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
body_font = Font(name="Arial", size=10)
thin = Side(style="thin", color="C9BFA8")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

for cell in ws[1]:
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = border

widths = [6, 14, 18, 55, 28, 14, 20, 40]
for i, w in enumerate(widths, start=1):
    ws.column_dimensions[get_column_letter(i)].width = w

for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=8):
    for cell in row:
        cell.font = body_font
        cell.alignment = Alignment(vertical="top", wrap_text=True)
        cell.border = border
    verify_cell = row[6]  # "재검증 필요 여부" 열
    if verify_cell.value and "권장" in str(verify_cell.value):
        verify_cell.fill = PatternFill(start_color="FCEEEA", end_color="FCEEEA", fill_type="solid")
    elif verify_cell.value and "완료" in str(verify_cell.value):
        verify_cell.fill = PatternFill(start_color="EAF2EA", end_color="EAF2EA", fill_type="solid")

ws.freeze_panes = "A2"
ws.auto_filter.ref = f"A1:H{ws.max_row}"

# ---- 3. 요약 시트 (CSV 행수에 맞춰 수식 범위 자동 조정) ----
ws2 = wb.create_sheet("요약")

# 분류별 개수는 데이터 기준으로 자동 계산
from collections import Counter
cat_counts = Counter(r[1] for r in rows if len(r) > 1)

summary = [
    ["항목", "값"],
    ["총 항목 수", len(rows)],
]
for cat, cnt in cat_counts.items():
    summary.append([cat, cnt])
summary += [
    ["", ""],
    ["재검증 권장 항목 수", f"=COUNTIF('사실관계 목록'!G2:G{last_row},\"*권장*\")"],
    ["재검증 불필요(널리 알려진 사실) 항목 수", f"=COUNTIF('사실관계 목록'!G2:G{last_row},\"불필요*\")"],
    ["", ""],
    ["우선순위 높은 재검증 대상", "최광옥, 김좌진(신민회 이력), 이범석, 박희도(숭실학교 졸업), 이종호(안중근 의거 연루)"],
    ["사유", "이 5개 항목은 서로 다른 계열(학교-단체, 계몽운동-무장투쟁, 독립운동-친일)을 잇는 핵심 연결고리로 온톨로지에서 사용되었기 때문에, 사실이 아닐 경우 그래프 구조 해석 자체가 바뀜"],
]
for r in summary:
    ws2.append(r)

for cell in ws2[1]:
    cell.fill = header_fill
    cell.font = header_font
for row in ws2.iter_rows(min_row=2, max_row=ws2.max_row, max_col=2):
    for cell in row:
        cell.font = body_font
        cell.alignment = Alignment(vertical="top", wrap_text=True)
ws2.column_dimensions["A"].width = 30
ws2.column_dimensions["B"].width = 70

wb.save(XLSX_PATH)
print(f"저장 완료: {XLSX_PATH} (데이터 {len(rows)}행)")
print("※ 수식 캐시 값을 채우려면 recalc.py를 실행하세요.")